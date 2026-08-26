"""Per-control working-paper Excel reports for SAP HANA DB ITGC."""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from src.reporting.excel_ole_embedder import embed_file_in_worksheet
except ImportError:  # pragma: no cover
    embed_file_in_worksheet = None  # type: ignore[assignment]

try:
    from core.findings_master_detail import display_control_id
except ImportError:  # pragma: no cover
    def display_control_id(control_id: str, catalog_entry: Optional[Dict[str, Any]] = None) -> str:
        entry = catalog_entry or {}
        ayalon = str(entry.get("control_id_ayalon", "") or "").strip()
        return ayalon or str(control_id or "").strip() or "-"

_logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_KEY_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_KEY_FONT = Font(bold=True, size=11)
_FINDING_FILL = PatternFill(start_color="FFFFE6E6", end_color="FFFFE6E6", fill_type="solid")
_THIN = Side(border_style="thin", color="FF808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)
_WRAP_RIGHT_RTL = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SYSTEM_NAME = "SAP HANA DB"
_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp"}


def _sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", str(name or "")).strip()
    return (cleaned or "Sheet")[:31]


def _set_rtl(sheet) -> None:
    sheet.sheet_view.rightToLeft = True


def _apply_header(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.border = _BORDER


def _apply_value_cell(cell, *, fill: PatternFill | None = None, font: Font | None = None) -> None:
    cell.alignment = _WRAP_RIGHT
    cell.border = _BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def write_control_working_paper(
    *,
    control_id: str,
    catalog_entry: Optional[Dict[str, Any]] = None,
    summary_record: Optional[Dict[str, Any]] = None,
    detail_rows: Optional[Sequence[Dict[str, Any]]] = None,
    raw_population_rows: Optional[Sequence[Dict[str, Any]]] = None,
    ipe_entries: Optional[Sequence[Dict[str, Any]]] = None,
    output_path: Path,
    notes: Optional[List[str]] = None,
    raw_population_note: Optional[str] = None,
    compensating_control_entry: Optional[Dict[str, Any]] = None,
) -> Path:
    """Build a working-paper Excel file (4+ sheets) and save to *output_path*."""
    catalog_entry = catalog_entry or {}
    summary_record = summary_record or {}
    detail_rows = list(detail_rows or [])
    raw_population_rows = list(raw_population_rows or [])
    ipe_entries = list(ipe_entries or [])
    notes = list(notes or [])

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    display_id = display_control_id(control_id, catalog_entry)
    overview.title = _sanitize_sheet_name(display_id)
    _write_overview_sheet(
        overview,
        control_id=control_id,
        catalog_entry=catalog_entry,
        summary_record=summary_record,
        notes=notes,
    )

    ipe_sheet = workbook.create_sheet("IPE")
    _write_ipe_sheet(ipe_sheet, ipe_entries)

    population_sheet = workbook.create_sheet("אוכלוסיה נבחנת")
    _write_population_sheet(
        population_sheet,
        raw_population_rows,
        detail_rows,
        note=raw_population_note,
    )

    findings_sheet = workbook.create_sheet("ריכוז ממצאים")
    _write_findings_sheet(findings_sheet, detail_rows)

    embed_pending = False
    if compensating_control_entry:
        comp_sheet = workbook.create_sheet("בקרה מפצה")
        embed_pending = _write_compensating_control_sheet(comp_sheet, compensating_control_entry)

    workbook.save(output_path)

    if compensating_control_entry and embed_pending and embed_file_in_worksheet is not None:
        stored_path = Path(str(compensating_control_entry.get("stored_path", "")))
        embedded_path = None
        if stored_path.exists():
            embedded_path = embed_file_in_worksheet(
                output_path,
                "בקרה מפצה",
                stored_path,
            )
            if embedded_path is None:
                _append_compensating_embed_failure_note(output_path, compensating_control_entry)
            else:
                output_path = Path(embedded_path)

    return output_path


def _write_overview_sheet(
    sheet,
    *,
    control_id: str,
    catalog_entry: Dict[str, Any],
    summary_record: Dict[str, Any],
    notes: List[str],
) -> None:
    _set_rtl(sheet)
    display_id = display_control_id(control_id, catalog_entry)
    rows = [
        ("מזהה בקרה", display_id),
        ("כותרת בקרה", catalog_entry.get("title_he") or summary_record.get("title_he") or "-"),
        ("שם מערכת", SYSTEM_NAME),
        ("תהליך", catalog_entry.get("process") or "-"),
        ("קטגוריה", catalog_entry.get("category") or catalog_entry.get("domain") or "-"),
        ("תיאור סיכון", catalog_entry.get("risk_description") or "-"),
        ("תיאור בקרה", catalog_entry.get("description") or summary_record.get("description") or "-"),
        ("סוג בדיקה", summary_record.get("check_type") or catalog_entry.get("analysis_type") or "-"),
        ("רמת סיכון (ממצאים)", summary_record.get("risk_level") or "-"),
        ("רשומות תקינות", summary_record.get("valid_records", "-")),
        ("רשומות עם ממצא", summary_record.get("finding_records", "-")),
        ('סה"כ רשומות', summary_record.get("total_records", "-")),
        ("קובץ מקור", summary_record.get("source_file") or "-"),
        ("תאריך הפקה", summary_record.get("extraction_date") or "-"),
        ("צעדי טסט", catalog_entry.get("test_steps_override") or "-"),
        ("הערות", " | ".join(notes) if notes else (catalog_entry.get("notes") or "-")),
    ]
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90
    for index, (key, value) in enumerate(rows, start=1):
        key_cell = sheet.cell(row=index, column=1, value=str(key))
        _apply_value_cell(key_cell, fill=_KEY_FILL, font=_KEY_FONT)
        value_cell = sheet.cell(row=index, column=2, value=str(value if value is not None else "-"))
        _apply_value_cell(value_cell)
        if key == "צעדי טסט":
            # Force RTL reading order so Hebrew multiline text aligns to the right.
            value_cell.alignment = _WRAP_RIGHT_RTL
            line_count = max(1, str(value).count("\n") + 1)
            sheet.row_dimensions[index].height = max(80, min(line_count * 16, 420))
        else:
            sheet.row_dimensions[index].height = 18 if index not in {7} else 80


def _write_ipe_sheet(sheet, ipe_entries: Sequence[Dict[str, Any]]) -> None:
    _set_rtl(sheet)
    headers = ["#", "שם קובץ מקורי", "תאריך הוספה", "נתיב שמור", "תצוגה"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        _apply_header(cell)
        sheet.column_dimensions[get_column_letter(col)].width = 24 if col < 5 else 40

    if not ipe_entries:
        empty = sheet.cell(row=2, column=1, value="לא נמצאו תיעודי IPE עבור בקרה זו.")
        empty.alignment = _WRAP_RIGHT
        return

    image_row = 2
    for index, entry in enumerate(ipe_entries, start=1):
        values = [
            index,
            entry.get("original_filename", "-"),
            entry.get("added_at", "-"),
            entry.get("stored_path", "-"),
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=image_row, column=col, value=value)
            _apply_value_cell(cell)

        stored = Path(str(entry.get("stored_path", "")))
        if stored.exists() and stored.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif"}:
            try:
                image = XLImage(str(stored))
                image.width = min(getattr(image, "width", 320) or 320, 320)
                image.height = min(getattr(image, "height", 180) or 180, 180)
                sheet.add_image(image, f"E{image_row}")
                sheet.row_dimensions[image_row].height = max(image.height * 0.75, 40)
            except Exception:
                sheet.cell(row=image_row, column=5, value="לא ניתן להטמיע תמונה")
        image_row += 1


def _write_population_sheet(
    sheet,
    raw_population_rows: Sequence[Dict[str, Any]],
    detail_rows: Sequence[Dict[str, Any]],
    *,
    note: Optional[str] = None,
) -> None:
    _set_rtl(sheet)
    start_row = 1
    if note:
        note_cell = sheet.cell(row=1, column=1, value=note)
        note_cell.font = Font(bold=True, color="FFC00000")
        note_cell.alignment = _WRAP_RIGHT
        start_row = 3

    if not raw_population_rows:
        empty = sheet.cell(row=start_row, column=1, value="אין אוכלוסייה גולמית זמינה לבקרה זו.")
        empty.alignment = _WRAP_RIGHT
        return

    finding_keys = set()
    for detail in detail_rows:
        for key in ("user_name", "actual_value", "title", "description"):
            value = str(detail.get(key) or "").strip()
            if value:
                finding_keys.add(value.casefold())

    headers = list(raw_population_rows[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col, value=str(header))
        _apply_header(cell)
        sheet.column_dimensions[get_column_letter(col)].width = 18

    for row_offset, row in enumerate(raw_population_rows, start=1):
        excel_row = start_row + row_offset
        row_text = " ".join(str(value) for value in row.values() if value is not None)
        is_finding = any(token in row_text.casefold() for token in finding_keys) if finding_keys else False
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=excel_row, column=col, value=row.get(header, ""))
            _apply_value_cell(cell, fill=_FINDING_FILL if is_finding else None)


def _write_findings_sheet(sheet, detail_rows: Sequence[Dict[str, Any]]) -> None:
    _set_rtl(sheet)
    headers = [
        "קטגוריה",
        "רמת סיכון",
        "תיאור",
        "סוג בדיקה",
        "ערך בפועל",
        "ערך מצופה",
        "סטטוס",
        "קובץ מקור",
        "תאריך הפקה",
        "תיאור מלא",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        _apply_header(cell)
        sheet.column_dimensions[get_column_letter(col)].width = 22

    if not detail_rows:
        empty = sheet.cell(row=2, column=1, value="אין ממצאים לבקרה זו.")
        empty.alignment = _WRAP_RIGHT
        return

    for row_index, detail in enumerate(detail_rows, start=2):
        values = [
            detail.get("category", "-"),
            detail.get("risk_level", "-"),
            detail.get("title", "-"),
            detail.get("comparison_rule") or detail.get("check_type") or "-",
            detail.get("actual_value", "-"),
            detail.get("expected_value", "-"),
            detail.get("status", "-"),
            detail.get("source_file") or detail.get("source_slot") or "-",
            detail.get("extract_date", "-"),
            detail.get("description", "-"),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value="" if value is None else str(value))
            fill = _FINDING_FILL if str(detail.get("status", "")) != "Compliant" else None
            _apply_value_cell(cell, fill=fill)


def _write_compensating_control_sheet(sheet, entry: Dict[str, Any]) -> bool:
    """Write a 'בקרה מפצה' sheet. Returns True when OLE embed is needed after save."""
    _set_rtl(sheet)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90

    meta_rows = [
        ("מזהה בקרה", entry.get("control_id", "-")),
        ("שם קובץ מקורי", entry.get("original_filename", "-")),
        ("תאריך הוספה", entry.get("added_at", "-")),
        ("נתיב שמור", entry.get("stored_path", "-")),
    ]
    for row_index, (key, value) in enumerate(meta_rows, start=1):
        key_cell = sheet.cell(row=row_index, column=1, value=str(key))
        _apply_value_cell(key_cell, fill=_KEY_FILL, font=_KEY_FONT)
        value_cell = sheet.cell(row=row_index, column=2, value=str(value if value is not None else "-"))
        _apply_value_cell(value_cell)
        sheet.row_dimensions[row_index].height = 18

    stored_path = Path(str(entry.get("stored_path", "")))
    if not stored_path.exists():
        note = sheet.cell(row=6, column=1, value="קובץ התיעוד לא נמצא בנתיב השמור.")
        note.alignment = _WRAP_RIGHT
        return False

    suffix = stored_path.suffix.lower()
    if suffix in _IMAGE_SUFFIXES:
        try:
            img = XLImage(str(stored_path))
            img.anchor = "A6"
            sheet.add_image(img)
        except Exception:
            note = sheet.cell(row=6, column=1, value=f"[לא ניתן להטמיע תמונה: {stored_path.name}]")
            note.alignment = _WRAP_RIGHT
        return False

    embed_label = sheet.cell(row=6, column=1, value="קובץ מוטמע")
    _apply_value_cell(embed_label, fill=_KEY_FILL, font=_KEY_FONT)
    pending_cell = sheet.cell(
        row=6,
        column=2,
        value="הקובץ יוטמע בתוך נייר העבודה לאחר השמירה.",
    )
    _apply_value_cell(pending_cell)
    sheet.row_dimensions[7].height = 280
    return True


def _append_compensating_embed_failure_note(output_path: Path, entry: Dict[str, Any]) -> None:
    """Best-effort note when OLE embedding could not be completed."""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(output_path)
        if "בקרה מפצה" not in workbook.sheetnames:
            return
        sheet = workbook["בקרה מפצה"]
        filename = str(entry.get("original_filename", "") or "")
        note_cell = sheet.cell(
            row=8,
            column=1,
            value=(
                f"לא ניתן היה להטמיע את הקובץ {filename} בתוך Excel. "
                f"הקובץ המקורי נשמר בנתיב: {entry.get('stored_path', '-')}"
            ),
        )
        note_cell.alignment = _WRAP_RIGHT
        sheet.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
        workbook.save(output_path)
    except Exception as exc:
        _logger.warning("Failed to append compensating embed failure note: %s", exc)


def safe_working_paper_filename(control_id: str, timestamp: str) -> str:
    safe_id = re.sub(r"[\\/*?:\[\]&]", "_", str(control_id or "control"))
    return f"{safe_id}_working_paper_{timestamp}.xlsx"
