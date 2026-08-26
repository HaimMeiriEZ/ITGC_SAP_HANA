"""Tests for IPE evidence storage and working-paper export."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image as PilImage

from core.ipe_evidence import (
    IpeEvidenceRepository,
    build_slot_to_controls_mapping,
    controls_for_slot,
)
from src.reporting.working_paper_report import write_control_working_paper


def test_ipe_repository_add_remove_and_reload(tmp_path: Path):
    output_dir = tmp_path / "output"
    base_dir = tmp_path
    repo = IpeEvidenceRepository(output_dir=output_dir, base_dir=base_dir)
    data: dict = {}

    source = tmp_path / "shot.png"
    PilImage.new("RGB", (20, 10), color=(255, 0, 0)).save(source)

    entry = repo.add_image("USERS", source, ["DB-AM-01_PLACEHOLDER"], data)
    assert entry["id"]
    assert Path(entry["stored_path"]).exists()
    assert "USERS" in data

    reloaded = repo.load()
    assert len(reloaded["USERS"]) == 1

    repo.remove_image("USERS", entry["id"], data)
    assert "USERS" not in data or data.get("USERS") == []
    assert not Path(entry["stored_path"]).exists()


def test_ipe_repository_clear_all_removes_files_and_metadata(tmp_path: Path):
    output_dir = tmp_path / "output"
    base_dir = tmp_path
    repo = IpeEvidenceRepository(output_dir=output_dir, base_dir=base_dir)
    data: dict = {}

    source = tmp_path / "shot.png"
    PilImage.new("RGB", (20, 10), color=(0, 255, 0)).save(source)

    entry = repo.add_image("USERS", source, ["DB-AM-01_PLACEHOLDER"], data)
    stored = Path(entry["stored_path"])
    assert stored.exists()
    assert repo.load().get("USERS")

    cleared = repo.clear_all(data)
    assert cleared == {}
    assert data == {}
    assert not stored.exists()
    assert repo.load() == {}
    evidence_root = base_dir / "data" / "evidence"
    if evidence_root.exists():
        remaining = [p for p in evidence_root.rglob("*") if p.is_file()]
        assert remaining == []


def test_slot_to_controls_mapping_inverts_catalog():
    catalog = {
        "DB-A": {"required_slots": ["USERS", "GRANTED_PRIVILEGES"]},
        "DB-B": {"required_slots": ["USERS"]},
    }
    mapping = build_slot_to_controls_mapping(catalog)
    assert set(mapping["USERS"]) == {"DB-A", "DB-B"}
    assert controls_for_slot("GRANTED_PRIVILEGES", catalog) == ["DB-A"]


def test_missing_ipe_gate_logic():
    # Lightweight mirror of GUI gate: loaded slots without evidence fail.
    loaded = {"USERS": object(), "M_PASSWORD_POLICY": object()}
    evidence = {"USERS": [{"id": "1"}]}
    missing = [
        slot
        for slot in loaded
        if not evidence.get(slot)
    ]
    assert missing == ["M_PASSWORD_POLICY"]


def test_alias_slots_use_primary_ipe_evidence():
    from core.ipe_evidence import collect_missing_ipe_slots, resolve_primary_ipe_slot

    assert resolve_primary_ipe_slot("AUDIT_LOG") == "AUDIT_TRAIL"
    assert resolve_primary_ipe_slot("EFFECTIVE_ROLES") == "GRANTED_ROLES"
    assert resolve_primary_ipe_slot(
        "EFFECTIVE_PRIVILEGE_GRANTEES",
        {"GRANTED_PRIVILEGES": "a.txt", "EFFECTIVE_PRIVILEGE_GRANTEES": "a.txt"},
    ) == "GRANTED_PRIVILEGES"
    assert resolve_primary_ipe_slot(
        "EFFECTIVE_PRIVILEGE_GRANTEES",
        {"EFFECTIVE_PRIVILEGE_GRANTEES": "own.txt"},
    ) == "EFFECTIVE_PRIVILEGE_GRANTEES"

    messages = collect_missing_ipe_slots(
        ["AUDIT_TRAIL", "AUDIT_LOG", "GRANTED_ROLES", "EFFECTIVE_ROLES"],
        {
            "AUDIT_TRAIL": [{"id": "1"}],
            "GRANTED_ROLES": [{"id": "2"}],
        },
        loaded_files={
            "AUDIT_TRAIL": "audit.txt",
            "AUDIT_LOG": "audit.txt",
            "GRANTED_ROLES": "roles.txt",
            "EFFECTIVE_ROLES": "roles.txt",
        },
        slot_labels={
            "AUDIT_TRAIL": "ראיות Audit בפועל",
            "GRANTED_ROLES": "הקצאות תפקידים",
        },
    )
    assert messages == []

    messages_missing = collect_missing_ipe_slots(
        ["AUDIT_TRAIL", "AUDIT_LOG"],
        {},
        loaded_files={"AUDIT_TRAIL": "audit.txt", "AUDIT_LOG": "audit.txt"},
        slot_labels={"AUDIT_TRAIL": "ראיות Audit בפועל"},
    )
    assert len(messages_missing) == 1
    assert "ראיות Audit בפועל" in messages_missing[0]
    assert "AUDIT_LOG" not in messages_missing[0]


def test_write_control_working_paper_creates_four_sheets(tmp_path: Path):
    image_path = tmp_path / "ipe.png"
    PilImage.new("RGB", (40, 20), color=(0, 128, 255)).save(image_path)
    output = tmp_path / "wp.xlsx"

    write_control_working_paper(
        control_id="DB-AM-04_PLACEHOLDER",
        catalog_entry={
            "control_id_ayalon": "DB-AM-01_04_79",
            "title_he": "הרשאות קריטיות",
            "process": "גישה",
            "description": "תיאור",
            "risk_description": "סיכון",
            "test_steps_override": "1. צעד ראשון\n2. צעד שני",
        },
        summary_record={
            "title_he": "הרשאות קריטיות",
            "check_type": "CRITICAL_PRIVILEGES",
            "risk_level": "High",
            "valid_records": 0,
            "finding_records": 1,
            "total_records": 1,
            "source_file": "priv.txt",
            "extraction_date": "2025-08-21",
            "description": "desc",
        },
        detail_rows=[
            {
                "category": "Access",
                "risk_level": "High",
                "title": "ממצא",
                "comparison_rule": "rule",
                "actual_value": "USER1",
                "expected_value": "-",
                "status": "Non-Compliant",
                "source_file": "priv.txt",
                "extract_date": "2025-08-21",
                "description": "full",
            }
        ],
        raw_population_rows=[{"USER_NAME": "USER1", "PRIVILEGE": "USER ADMIN"}],
        ipe_entries=[
            {
                "original_filename": "ipe.png",
                "stored_path": str(image_path),
                "added_at": "2025-08-21T10:00:00",
                "control_ids": ["DB-AM-04_PLACEHOLDER"],
            }
        ],
        output_path=output,
    )

    assert output.exists()
    workbook = load_workbook(output)
    assert len(workbook.sheetnames) == 4
    assert "IPE" in workbook.sheetnames
    assert "אוכלוסיה נבחנת" in workbook.sheetnames
    assert "ריכוז ממצאים" in workbook.sheetnames

    overview = workbook[workbook.sheetnames[0]]
    assert overview.title == "DB-AM-01_04_79"
    assert overview["A1"].value == "מזהה בקרה"
    assert overview["B1"].value == "DB-AM-01_04_79"
    assert overview["A15"].value == "צעדי טסט"
    steps_alignment = overview["B15"].alignment
    assert steps_alignment.horizontal == "right"
    assert steps_alignment.wrap_text is True
    assert getattr(steps_alignment, "readingOrder", None) == 2
