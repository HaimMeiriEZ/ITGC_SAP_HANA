"""Import filled Excel catalog sheet into controls_catalog.json."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "knowledge_base" / "_templates" / "controls_catalog_fill_template.xlsx"
CATALOG_PATH = ROOT / "data" / "knowledge_base" / "controls_catalog.json"
REVIEW_PATH = ROOT / "data" / "knowledge_base" / "_templates" / "_review_catalog.json"
SLOTS_PATH = ROOT / "data" / "knowledge_base" / "slot_definitions.json"

SHEET_CATALOG = "קטלוג בקרות"

SLOT_ALIASES = {
    "M_AUDIT_POLICIES": "AUDIT_POLICIES",
    "M_AUDIT_LOG": "AUDIT_TRAIL",
}

# Canonical slots when Excel still uses view names or is incomplete.
REQUIRED_SLOTS_OVERRIDES: dict[str, list[str]] = {
    "DB-AL-01_PLACEHOLDER": ["M_INIFILE_CONTENTS", "AUDIT_POLICIES", "AUDIT_TRAIL"],
    "DB-UAR-01_PLACEHOLDER": ["USERS", "GRANTED_ROLES", "EFFECTIVE_PRIVILEGE_GRANTEES"],
}

VALIDATOR_REF_BY_ANALYSIS_TYPE: dict[str, str] = {
    "USER_POPULATION_REVIEW": "validate_user_population",
    "PASSWORD_POLICY_BASELINE": "validate_password_policy_baseline",
    "PASSWORD_POLICY_EXEMPTIONS": "validate_password_policy_exemptions",
    "AUDIT_POLICY_ENABLED": "validate_audit_policy_enabled",
    "USER_PROVISIONING": "validate_user_provisioning",
    "BUILTIN_USERS_LOCKDOWN": "validate_builtin_users_lockdown",
    "SYSTEM_USER_LOCKDOWN": "validate_system_user_lockdown",
    "CRITICAL_PRIVILEGES": "validate_critical_privileges",
    "PERIODIC_UAR": "validate_periodic_uar",
    "ANALYTIC_OBJECT_ACCESS": "validate_analytic_object_access",
    "SYSTEM_TABLE_ACCESS": "validate_system_table_access",
}

EXCEL_TO_JSON = {
    "control_id (placeholder)": "control_id",
    "control_id_ayalon (למילוי)": "control_id_ayalon",
    "title_he": "title_he",
    "domain": "domain",
    "category": "category",
    "required_slots": "required_slots",
    "in_scope": "in_scope",
    "implementation_status": "implementation_status",
    "validator_ref": "validator_ref",
    "analysis_type": "analysis_type",
    "process": "process",
    "description (למילוי)": "description",
    "risk_description (למילוי)": "risk_description",
    "test_steps_override (למילוי)": "test_steps_override",
    "notes": "notes",
}


def parse_slots(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).replace(";", ",")
    parts = [part.strip() for part in text.split(",") if part.strip()]
    normalized: list[str] = []
    for part in parts:
        normalized.append(SLOT_ALIASES.get(part, part))
    return normalized


def is_filled(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text.upper() not in {"NONE", "NULL", "(ריק)"}


def read_catalog_rows(ws) -> tuple[list[str], list[dict]]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        if all(not is_filled(value) for value in values):
            continue
        record = {headers[i]: values[i] for i in range(len(headers))}
        record["_excel_row"] = row_idx
        rows.append(record)
    return headers, rows


def to_bool(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    return text not in {"false", "0", "no", "לא"}


def build_catalog_entry(row: dict, valid_slots: set[str]) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    control_id = str(row.get("control_id (placeholder)", "")).strip()
    slots = parse_slots(row.get("required_slots"))
    if control_id in REQUIRED_SLOTS_OVERRIDES:
        canonical = REQUIRED_SLOTS_OVERRIDES[control_id]
        if slots != canonical:
            warnings.append(
                f"{control_id}: required_slots normalized {slots!r} -> {canonical!r}"
            )
            slots = list(canonical)

    unknown = [slot for slot in slots if slot not in valid_slots]
    if unknown:
        warnings.append(f"{control_id}: unknown slot keys {unknown!r}")

    analysis_type = str(row.get("analysis_type") or "").strip()
    validator_ref = row.get("validator_ref")
    if not is_filled(validator_ref) and analysis_type:
        validator_ref = VALIDATOR_REF_BY_ANALYSIS_TYPE.get(analysis_type)

    entry = {
        "control_id": control_id,
        "control_id_ayalon": str(row.get("control_id_ayalon (למילוי)") or "").strip(),
        "title_he": str(row.get("title_he") or "").strip(),
        "domain": str(row.get("domain") or "").strip(),
        "category": str(row.get("category") or "").strip(),
        "required_slots": slots,
        "implementation_status": str(row.get("implementation_status") or "placeholder").strip(),
        "validator_ref": str(validator_ref).strip() if is_filled(validator_ref) else None,
        "in_scope": to_bool(row.get("in_scope")),
        "analysis_type": analysis_type,
        "description": str(row.get("description (למילוי)") or "").strip(),
        "process": str(row.get("process") or "").strip(),
        "risk_description": str(row.get("risk_description (למילוי)") or "").strip(),
        "test_steps_override": str(row.get("test_steps_override (למילוי)") or "").strip(),
        "notes": str(row.get("notes") or "").strip(),
    }
    return entry, warnings


def patch_excel_slots(ws, headers: list, rows: list[dict], entries: list[dict]) -> None:
    try:
        slots_col = headers.index("required_slots") + 1
    except ValueError:
        return
    entry_by_id = {entry["control_id"]: entry for entry in entries}
    for row in rows:
        control_id = str(row.get("control_id (placeholder)", "")).strip()
        entry = entry_by_id.get(control_id)
        if not entry:
            continue
        ws.cell(row["_excel_row"], slots_col, ", ".join(entry["required_slots"]))


def main() -> None:
    slots = json.loads(SLOTS_PATH.read_text(encoding="utf-8"))["slots"]
    valid_slots = set(slots)

    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_CATALOG]
    headers, rows = read_catalog_rows(ws)

    entries: list[dict] = []
    all_warnings: list[str] = []
    for row in rows:
        entry, warnings = build_catalog_entry(row, valid_slots)
        entries.append(entry)
        all_warnings.extend(warnings)

    patch_excel_slots(ws, headers, rows, entries)
    wb.save(XLSX_PATH)

    payload = {
        "_schema_version": 2,
        "_description": "קטלוג בקרות ITGC ל-SAP HANA DB — יובא מ-controls_catalog_fill_template.xlsx",
        "_source": str(XLSX_PATH.relative_to(ROOT)).replace("\\", "/"),
        "controls": entries,
    }
    CATALOG_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    review = {
        "count": len(entries),
        "controls": [{k: row.get(k) for k in row if k != "_excel_row"} for row in rows],
        "import_warnings": all_warnings,
        "validator_ref_plan": VALIDATOR_REF_BY_ANALYSIS_TYPE,
    }
    REVIEW_PATH.write_text(json.dumps(review, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Imported {len(entries)} controls -> {CATALOG_PATH}")
    if all_warnings:
        print("Warnings:")
        for warning in all_warnings:
            print(f"  - {warning}")
    else:
        print("No slot normalization warnings.")


if __name__ == "__main__":
    main()
