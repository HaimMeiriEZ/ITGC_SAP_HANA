"""Sync Excel dependent sheets from the filled 'קטלוג בקרות' sheet."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "knowledge_base" / "_templates" / "controls_catalog_fill_template.xlsx"
SLOTS_PATH = ROOT / "data" / "knowledge_base" / "slot_definitions.json"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HINT_FILL = PatternFill("solid", fgColor="FFF2CC")
EMPTY_FILL = PatternFill("solid", fgColor="FCE4D6")
FILLED_FILL = PatternFill("solid", fgColor="E2EFDA")
REF_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top", readingOrder=2)
RTL = Alignment(wrap_text=True, vertical="center", readingOrder=2)

SHEET_CATALOG = "קטלוג בקרות"
SHEET_MAP = "מיפוי בקרה-טבלאות"
SHEET_SLOTS = "הגדרות טבלאות"
SHEET_GAPS = "שדות להשלמה"
SHEET_INSTRUCTIONS = "הנחיות"


def style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def clear_sheet(ws: Worksheet) -> None:
    ws.delete_rows(1, ws.max_row)
    for dv in list(ws.data_validations.dataValidation):
        ws.data_validations.dataValidation.remove(dv)
    ws.auto_filter.ref = None


def parse_slots(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).replace(";", ",")
    return [part.strip() for part in text.split(",") if part.strip()]


def is_filled(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text.upper() not in {"NONE", "NULL", "(ריק)"}


def format_required_columns(slot_def: dict) -> str:
    req_cols = ", ".join(slot_def.get("required_columns") or [])
    any_groups = slot_def.get("required_any_groups") or []
    if any_groups:
        extra = " | OR: " + " ; ".join(["/".join(group) for group in any_groups])
        return (req_cols + extra) if req_cols else extra.lstrip(" | ")
    return req_cols


def read_catalog(ws: Worksheet) -> tuple[list[str], list[dict]]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        if all(not is_filled(value) for value in values):
            continue
        record = {headers[i]: values[i] for i in range(len(headers))}
        record["_excel_row"] = row_idx
        rows.append(record)
    return headers, rows


def catalog_status(control: dict) -> str:
    missing: list[str] = []
    if not is_filled(control.get("control_id_ayalon (למילוי)")):
        missing.append("control_id_ayalon")
    if not is_filled(control.get("description (למילוי)")):
        missing.append("description")
    if not is_filled(control.get("risk_description (למילוי)")):
        missing.append("risk_description")
    if not is_filled(control.get("test_steps_override (למילוי)")):
        missing.append("test_steps")
    optional: list[str] = []
    if not is_filled(control.get("validator_ref")):
        optional.append("validator_ref")
    if not is_filled(control.get("analysis_type")):
        optional.append("analysis_type")
    if missing:
        return "חסר: " + ", ".join(missing)
    if optional:
        return "מולא — ממתין ל: " + ", ".join(optional)
    return "מולא במלואו"


def update_catalog_status(ws: Worksheet, headers: list, controls: list[dict]) -> None:
    try:
        status_col = headers.index("סטטוס מילוי") + 1
    except ValueError:
        return
    for control in controls:
        cell = ws.cell(control["_excel_row"], status_col, catalog_status(control))
        cell.fill = HINT_FILL
        cell.alignment = WRAP
        cell.border = THIN


def rebuild_mapping(ws: Worksheet, controls: list[dict], slots: dict) -> None:
    clear_sheet(ws)
    ws.sheet_view.rightToLeft = True
    headers = [
        "control_id",
        "control_id_ayalon",
        "title_he",
        "domain",
        "slot_key (טבלה)",
        "שם טבלה בעברית",
        "רמת דרישה (חובה/מומלץ)",
        "עמודות חובה בטבלה",
        "הערות / שימוש בבקרה (למילוי)",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    row = 2
    for control in controls:
        placeholder = control.get("control_id (placeholder)") or ""
        ayalon = control.get("control_id_ayalon (למילוי)") or ""
        title = control.get("title_he") or ""
        domain = control.get("domain") or ""
        slot_keys = parse_slots(control.get("required_slots"))
        if not slot_keys:
            slot_keys = ["(לא הוגדר)"]
        for slot_key in slot_keys:
            slot_def = slots.get(slot_key, {})
            values = [
                placeholder,
                ayalon,
                title,
                domain,
                slot_key,
                slot_def.get("label_he", ""),
                "חובה",
                format_required_columns(slot_def) if slot_def else "",
                "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row, col, value)
                cell.border = THIN
                cell.alignment = WRAP
                cell.fill = EMPTY_FILL if col == 9 else FILLED_FILL
            ws.row_dimensions[row].height = 30
            row += 1

    autosize(ws, [28, 24, 28, 12, 36, 42, 18, 55, 40])
    ws.freeze_panes = "A2"
    last = max(row - 1, 1)
    ws.auto_filter.ref = f"A1:I{last}"
    dv_req = DataValidation(type="list", formula1='"חובה,מומלץ,תומך בקרה"', allow_blank=True)
    ws.add_data_validation(dv_req)
    dv_req.add(f"G2:G{last}")


def rebuild_slots(ws: Worksheet, controls: list[dict], slots: dict) -> None:
    clear_sheet(ws)
    ws.sheet_view.rightToLeft = True
    headers = [
        "slot_key",
        "שם בעברית",
        "דפוסי שם קובץ",
        "עמודות חובה",
        "קבוצות OR (לפחות אחת מכל קבוצה)",
        "עמודות אופציונליות",
        "בקרות שמשתמשות בטבלה זו",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    slot_to_controls: dict[str, list[str]] = {key: [] for key in slots}
    for control in controls:
        label = control.get("control_id_ayalon (למילוי)") or control.get("control_id (placeholder)") or ""
        title = control.get("title_he") or ""
        display = f"{label} ({title})" if title else str(label)
        for slot_key in parse_slots(control.get("required_slots")):
            slot_to_controls.setdefault(slot_key, []).append(display)

    row_idx = 2
    for slot_key, slot_def in slots.items():
        any_groups = slot_def.get("required_any_groups") or []
        any_txt = " ; ".join([" OR ".join(group) for group in any_groups]) if any_groups else ""
        values = [
            slot_key,
            slot_def.get("label_he", ""),
            ", ".join(slot_def.get("filename_patterns") or []),
            ", ".join(slot_def.get("required_columns") or []),
            any_txt,
            ", ".join(slot_def.get("optional_columns") or []),
            ", ".join(slot_to_controls.get(slot_key, [])),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = REF_FILL
        ws.row_dimensions[row_idx].height = 40
        row_idx += 1

    autosize(ws, [36, 46, 48, 32, 55, 48, 70])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row_idx - 1}"


def rebuild_gaps(ws: Worksheet, controls: list[dict]) -> None:
    clear_sheet(ws)
    ws.sheet_view.rightToLeft = True
    headers = ["control_id", "control_id_ayalon", "title_he", "שדה חסר", "ערך נוכחי", "ערך מוצע / למילוי", "הערות"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    gap_fields = [
        ("control_id_ayalon", "control_id_ayalon (למילוי)", "מזהה Ayalon סופי"),
        ("description", "description (למילוי)", "תיאור הבקרה מ-workpaper"),
        ("risk_description", "risk_description (למילוי)", "תיאור הסיכון"),
        ("test_steps_override", "test_steps_override (למילוי)", "צעדי בדיקה"),
        ("analysis_type", "analysis_type", "סוג ניתוח (אופציונלי)"),
        ("validator_ref", "validator_ref", "שם validator לאחר יישום"),
        ("notes", "notes", "הערות / מיפוי EY"),
    ]

    row = 2
    for control in controls:
        for field_label, json_key, hint in gap_fields:
            current = control.get(json_key)
            if is_filled(current):
                continue
            values = [
                control.get("control_id (placeholder)") or "",
                control.get("control_id_ayalon (למילוי)") or "",
                control.get("title_he") or "",
                field_label,
                "(ריק)",
                "",
                hint,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row, col, value)
                cell.border = THIN
                cell.alignment = WRAP
                if col == 6:
                    cell.fill = EMPTY_FILL
                elif col == 7:
                    cell.fill = HINT_FILL
                else:
                    cell.fill = FILLED_FILL
            ws.row_dimensions[row].height = 28
            row += 1

    if row == 2:
        ws.cell(2, 1, "אין שדות חובה חסרים — נותרו רק שדות אופציונליים אם יופיעו בהרצה הבאה.")
        ws.merge_cells("A2:G2")

    autosize(ws, [28, 24, 28, 22, 12, 40, 45])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(row - 1, 2)}"


def update_instructions(ws: Worksheet, control_count: int, mapping_count: int) -> None:
    ws["B4"] = (
        "שורה לכל בקרה (מקור אמת). שאר הגיליונות מסונכרנים ממנו. "
        f"מצבה נוכחית: {control_count} בקרות."
    )
    ws["B5"] = (
        "איזו טבלת HANA (slot) נדרשת לכל בקרה לפי עמודת required_slots בקטלוג. "
        f"כרגע {mapping_count} שורות מיפוי."
    )
    ws["B6"] = (
        "עמודות חובה/אופציונליות לכל slot מהמערכת (slot_definitions.json), "
        "כולל EFFECTIVE_PRIVILEGE_GRANTEES ו-CONFIGURATION_PARAMETER_PROPERTIES."
    )
    ws["B7"] = "רשימת פערים שנותרו בפועל אחרי מילוי קטלוג הבקרות."
    ws["B8"] = (
        "1) עדכנו את גיליון קטלוג בקרות\n"
        "2) הריצו: python scripts/sync_controls_fill_excel_sheets.py\n"
        "3) לאחר אישור — ייבוא ל-controls_catalog.json"
    )
    for row in range(4, 9):
        ws.cell(row, 2).alignment = WRAP
        ws.cell(row, 2).border = THIN


def main() -> None:
    slots = json.loads(SLOTS_PATH.read_text(encoding="utf-8"))["slots"]
    wb = load_workbook(XLSX_PATH)
    catalog_ws = wb[SHEET_CATALOG]
    headers, controls = read_catalog(catalog_ws)
    mapping_count = sum(max(len(parse_slots(c.get("required_slots"))), 1) for c in controls)

    update_catalog_status(catalog_ws, headers, controls)
    rebuild_mapping(wb[SHEET_MAP], controls, slots)
    rebuild_slots(wb[SHEET_SLOTS], controls, slots)
    rebuild_gaps(wb[SHEET_GAPS], controls)
    if SHEET_INSTRUCTIONS in wb.sheetnames:
        update_instructions(wb[SHEET_INSTRUCTIONS], len(controls), mapping_count)

    wb.save(XLSX_PATH)
    print(f"Updated: {XLSX_PATH}")
    print(f"Controls: {len(controls)}")
    print(f"Mapping rows: {mapping_count}")
    print(f"Slots: {len(slots)}")


if __name__ == "__main__":
    main()
