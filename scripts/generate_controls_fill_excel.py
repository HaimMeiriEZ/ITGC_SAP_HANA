"""Generate Excel templates for filling planned ITGC controls catalog details."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "knowledge_base" / "_templates" / "controls_catalog_fill_template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HINT_FILL = PatternFill("solid", fgColor="FFF2CC")
EMPTY_FILL = PatternFill("solid", fgColor="FCE4D6")
FILLED_FILL = PatternFill("solid", fgColor="E2EFDA")
REF_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top", readingOrder=2)
RTL = Alignment(wrap_text=True, vertical="center", readingOrder=2)

EY_MAP = {
    "DB-AM-01_PLACEHOLDER": "שימוש במשתמשים קריטיים",
    "DB-AM-02_PLACEHOLDER": "הרשאות ניהול ישירות למשתמשים",
    "DB-AM-03_PLACEHOLDER": "הרשאות רגישות דרך Role inheritance",
    "DB-PP-01_PLACEHOLDER": "מדיניות סיסמאות בטבלת HANA",
    "DB-PP-02_PLACEHOLDER": "פרמטרי סיסמאות ברמת INI",
    "DB-AL-01_PLACEHOLDER": "קיום והפעלה של Audit Policies",
    "DB-AL-02_PLACEHOLDER": "ראיות Audit בפועל מתקופת הביקורת",
    "DB-CF-01_PLACEHOLDER": "Configuration Hardening (INI)",
    "DB-UAR-01_PLACEHOLDER": "סקירת משתמשים, חריגים, ואישור מנהל",
}

REQ_LEVEL = {
    "USERS": "חובה",
    "M_PASSWORD_POLICY": "חובה",
    "GRANTED_PRIVILEGES": "חובה",
    "AUDIT_POLICIES": "חובה",
    "M_INIFILE_CONTENTS": "חובה",
    "GRANTED_ROLES": "מומלץ",
    "AUDIT_TRAIL": "מומלץ",
    "EFFECTIVE_PRIVILEGE_GRANTEES": "מומלץ",
    "CONFIGURATION_PARAMETER_PROPERTIES": "מומלץ",
}


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def autosize(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "הנחיות"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "תבנית מילוי קטלוג בקרות ITGC — SAP HANA DB"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")

    instructions = [
        ("מטרה", "מילוי פרטים חסרים לבקרות המתוכננות (Phase 1 placeholders) לפני החלפה במזהי Ayalon סופיים."),
        ("גיליון: קטלוג בקרות", "שורה לכל בקרה. עמודות כתומות = למילוי. עמודות ירוקות = ערך קיים מהמערכת."),
        ("גיליון: מיפוי בקרה-טבלאות", "איזו טבלת HANA (slot) נדרשת לכל בקרה, ורמת חובה."),
        ("גיליון: הגדרות טבלאות", "עמודות חובה/אופציונליות לכל slot — לעיון בלבד."),
        ("גיליון: שדות להשלמה", "רשימת פערים — מה ריק כרגע ודורש מילוי."),
        (
            "סדר עבודה",
            "1) מלאו מזהה Ayalon + תיאור + סיכון + צעדי בדיקה בגיליון קטלוג בקרות\n"
            "2) אשרו/עדכנו required_slots בגיליון מיפוי\n"
            "3) העתיקו ל-controls_catalog.json לפי CONTROLS_CATALOG_FILL_GUIDE.md",
        ),
        ("ערכי domain מותרים", "Access | Password | Audit | Config | UAR"),
        ("ערכי implementation_status", "placeholder | implemented"),
        ("ערכי in_scope", "TRUE | FALSE"),
        (
            "קבצים קשורים",
            "data/knowledge_base/controls_catalog.json\n"
            "docs/CONTROLS_CATALOG_FILL_GUIDE.md\n"
            "data/knowledge_base/slot_definitions.json",
        ),
    ]
    ws["A3"] = "נושא"
    ws["B3"] = "פירוט"
    style_header(ws, 2)
    ws.cell(3, 1).fill = HEADER_FILL
    ws.cell(3, 2).fill = HEADER_FILL
    for i, (key, value) in enumerate(instructions, 4):
        ws.cell(i, 1, key).border = THIN
        ws.cell(i, 1).fill = HINT_FILL
        ws.cell(i, 1).alignment = RTL
        ws.cell(i, 2, value).border = THIN
        ws.cell(i, 2).alignment = WRAP
        ws.row_dimensions[i].height = 45 if i == 9 else 35
    autosize(ws, [28, 85])
    ws.row_dimensions[1].height = 24


def build_catalog_sheet(wb: Workbook, controls: list[dict]) -> None:
    ws = wb.create_sheet("קטלוג בקרות")
    ws.sheet_view.rightToLeft = True
    headers = [
        "control_id (placeholder)",
        "control_id_ayalon (למילוי)",
        "title_he",
        "domain",
        "category",
        "required_slots",
        "in_scope",
        "implementation_status",
        "validator_ref",
        "analysis_type",
        "process",
        "description (למילוי)",
        "risk_description (למילוי)",
        "test_steps_override (למילוי)",
        "notes",
        "מיפוי EY",
        "סטטוס מילוי",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    fill_cols = {2, 12, 13, 14}
    pref_cols = {1, 3, 4, 5, 6, 7, 8, 10, 11, 15, 16}

    for row_idx, control in enumerate(controls, 2):
        slots_str = ", ".join(control.get("required_slots") or [])
        desc = control.get("description") or ""
        risk = control.get("risk_description") or ""
        steps = control.get("test_steps_override") or ""
        missing = ["control_id_ayalon"]
        if not desc:
            missing.append("description")
        if not risk:
            missing.append("risk_description")
        if not steps:
            missing.append("test_steps")
        status = "חסר: " + ", ".join(missing)

        values = [
            control["control_id"],
            "",
            control.get("title_he", ""),
            control.get("domain", ""),
            control.get("category", ""),
            slots_str,
            "TRUE" if control.get("in_scope", True) else "FALSE",
            control.get("implementation_status", "placeholder"),
            control.get("validator_ref") or "",
            control.get("analysis_type") or "",
            control.get("process") or "",
            desc,
            risk,
            steps,
            control.get("notes") or "",
            EY_MAP.get(control["control_id"], ""),
            status,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = THIN
            cell.alignment = WRAP
            if col in fill_cols:
                cell.fill = EMPTY_FILL
            elif col in pref_cols:
                cell.fill = FILLED_FILL
            elif col == 17:
                cell.fill = HINT_FILL
        ws.row_dimensions[row_idx].height = 40

    dv_domain = DataValidation(type="list", formula1='"Access,Password,Audit,Config,UAR"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"placeholder,implemented"', allow_blank=True)
    dv_scope = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_domain)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_scope)
    dv_domain.add("D2:D50")
    dv_status.add("H2:H50")
    dv_scope.add("G2:G50")

    autosize(ws, [28, 26, 22, 12, 24, 36, 10, 18, 22, 16, 32, 40, 40, 40, 40, 36, 36])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:Q{1 + len(controls)}"


def build_mapping_sheet(wb: Workbook, controls: list[dict], slots: dict) -> None:
    ws = wb.create_sheet("מיפוי בקרה-טבלאות")
    ws.sheet_view.rightToLeft = True
    headers = [
        "control_id",
        "title_he",
        "domain",
        "slot_key (טבלה)",
        "שם טבלה בעברית",
        "רמת דרישה (חובה/מומלץ)",
        "עמודות חובה בטבלה",
        "הערות / שימוש בבקרה (למילוי)",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    row = 2
    for control in controls:
        for slot in control.get("required_slots") or []:
            slot_def = slots.get(slot, {})
            req_cols = ", ".join(slot_def.get("required_columns") or [])
            any_groups = slot_def.get("required_any_groups") or []
            if any_groups:
                extra = " | OR: " + " ; ".join(["/".join(group) for group in any_groups])
                req_cols = (req_cols + extra) if req_cols else extra.lstrip(" | ")
            values = [
                control["control_id"],
                control.get("title_he", ""),
                control.get("domain", ""),
                slot,
                slot_def.get("label_he", ""),
                REQ_LEVEL.get(slot, "לבדיקה"),
                req_cols,
                "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row, col, value)
                cell.border = THIN
                cell.alignment = WRAP
                cell.fill = EMPTY_FILL if col == 8 else FILLED_FILL
            ws.row_dimensions[row].height = 30
            row += 1

    autosize(ws, [28, 24, 12, 24, 36, 18, 50, 45])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row - 1}"
    dv_req = DataValidation(type="list", formula1='"חובה,מומלץ,תומך בקרה"', allow_blank=True)
    ws.add_data_validation(dv_req)
    dv_req.add(f"F2:F{row}")


def build_slots_sheet(wb: Workbook, controls: list[dict], slots: dict) -> None:
    ws = wb.create_sheet("הגדרות טבלאות")
    ws.sheet_view.rightToLeft = True
    headers = [
        "slot_key",
        "שם בעברית",
        "דפוסי שם קובץ",
        "עמודות חובה",
        "קבוצות OR (לפחות אחת מכל קבוצה)",
        "עמודות אופציונליות",
        "בקרות שמשתמשות בטבלה זו",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    slot_to_controls: dict[str, list[str]] = {}
    for control in controls:
        for slot in control.get("required_slots") or []:
            slot_to_controls.setdefault(slot, []).append(control["control_id"])

    for row_idx, (slot_key, slot_def) in enumerate(slots.items(), 2):
        any_groups = slot_def.get("required_any_groups") or []
        any_txt = " ; ".join([" OR ".join(group) for group in any_groups]) if any_groups else ""
        values = [
            slot_key,
            slot_def.get("label_he", ""),
            ", ".join(slot_def.get("filename_patterns") or []),
            ", ".join(slot_def.get("required_columns") or []),
            any_txt,
            ", ".join(slot_def.get("optional_columns") or []),
            ", ".join(slot_to_controls.get(slot_key, [])),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = REF_FILL
        ws.row_dimensions[row_idx].height = 35

    autosize(ws, [22, 36, 40, 30, 50, 40, 50])
    ws.freeze_panes = "A2"


def build_gaps_sheet(wb: Workbook, controls: list[dict]) -> None:
    ws = wb.create_sheet("שדות להשלמה")
    ws.sheet_view.rightToLeft = True
    headers = ["control_id", "title_he", "שדה חסר", "ערך נוכחי", "ערך מוצע / למילוי", "הערות"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header(ws, len(headers))

    gap_fields = [
        ("control_id_ayalon", None, "מזהה Ayalon סופי (החלפת PLACEHOLDER)"),
        ("description", "description", "תיאור הבקרה מ-workpaper"),
        ("risk_description", "risk_description", "תיאור הסיכון"),
        ("test_steps_override", "test_steps_override", "צעדי בדיקה"),
        ("analysis_type", "analysis_type", "סוג ניתוח (אופציונלי)"),
        ("validator_ref", "validator_ref", "שם validator לאחר יישום"),
    ]

    row = 2
    for control in controls:
        for field_label, json_key, hint in gap_fields:
            current = "" if json_key is None else (control.get(json_key) or "")
            if json_key is None or not str(current).strip():
                values = [
                    control["control_id"],
                    control.get("title_he", ""),
                    field_label,
                    current or "(ריק)",
                    "",
                    hint,
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row, col, value)
                    cell.border = THIN
                    cell.alignment = WRAP
                    if col == 5:
                        cell.fill = EMPTY_FILL
                    elif col == 6:
                        cell.fill = HINT_FILL
                    else:
                        cell.fill = FILLED_FILL
                ws.row_dimensions[row].height = 28
                row += 1

    autosize(ws, [28, 24, 22, 14, 40, 45])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{row - 1}"


def build_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("מקרא צבעים")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "צבע"
    ws["B1"] = "משמעות"
    style_header(ws, 2)
    legend = [
        (EMPTY_FILL, "כתום — שדה למילוי / השלמה"),
        (FILLED_FILL, "ירוק — ערך קיים מהמערכת (ניתן לעדכן)"),
        (HINT_FILL, "צהוב — הנחיה / סטטוס"),
        (REF_FILL, "כחול — עיון בלבד (reference)"),
    ]
    for idx, (fill, text) in enumerate(legend, 2):
        ws.cell(idx, 1, "").fill = fill
        ws.cell(idx, 1).border = THIN
        ws.cell(idx, 2, text).border = THIN
        ws.cell(idx, 2).alignment = RTL
    autosize(ws, [12, 50])


def main() -> None:
    catalog = json.loads((ROOT / "data" / "knowledge_base" / "controls_catalog.json").read_text(encoding="utf-8"))
    slots = json.loads((ROOT / "data" / "knowledge_base" / "slot_definitions.json").read_text(encoding="utf-8"))["slots"]
    controls = catalog["controls"]

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_instructions(wb)
    build_catalog_sheet(wb, controls)
    build_mapping_sheet(wb, controls, slots)
    build_slots_sheet(wb, controls, slots)
    build_gaps_sheet(wb, controls)
    build_legend_sheet(wb)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Controls: {len(controls)}")
    print(f"Slots: {len(slots)}")


if __name__ == "__main__":
    main()
