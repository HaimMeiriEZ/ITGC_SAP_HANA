"""Export RTL Excel catalog of HANA DB input slots and analysis fields.

Mirrors the APP project workbook structure:
  - Sheet «ריכוז לפי טבלה»: summary per slot
  - Sheet «פירוט שדות»: field-level detail
  - Optional sheet «להעתקה»: three columns for easy paste

Output: docs/קבצי_קלט_ושדות.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "docs" / "קבצי_קלט_ושדות.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=11)
ALT_FILL = PatternFill("solid", fgColor="EEF3FB")
THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
HEADER_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

# Summary rows: table, required (כן/לא), description, expected file, relevant fields text
SUMMARY_ROWS: list[tuple[str, str, str, str, str]] = [
    (
        "USERS",
        "כן",
        "משתמשים — מקור חובה לסקירת גישה, סטטוס ותאריכי התחברות.",
        "users_export.csv",
        "חובה: USER_NAME, ו-LAST_SUCCESSFUL_CONNECT או LAST_SUCCESSFUL_CONNECT_DATE. "
        "לניתוח גם: USER_DEACTIVATED, CREATE_TIME, CREATOR, VALID_UNTIL, ADMIN_GIVEN_PASSWORD, "
        "IS_PASSWORD_LIFETIME_CHECK_ENABLED, LAST_PASSWORD_CHANGE_TIME, IS_RESTRICTED, USERGROUP_NAME",
    ),
    (
        "M_PASSWORD_POLICY",
        "כן",
        "מדיניות סיסמאות — פרמטרי אבטחת סיסמה ברמת המערכת.",
        "password_policy.csv",
        "חובה: PROPERTY, VALUE",
    ),
    (
        "GRANTED_PRIVILEGES",
        "כן",
        "הרשאות ישירות — מיפוי הרשאות למשתמשים/תפקידים.",
        "privileges.csv",
        "חובה: GRANTEE, PRIVILEGE. לניתוח: OBJECT_TYPE, GRANTEE_TYPE, SCHEMA_NAME, "
        "OBJECT_NAME, IS_GRANTABLE, IS_VALID",
    ),
    (
        "EFFECTIVE_PRIVILEGE_GRANTEES",
        "לא",
        "הרשאות אפקטיביות (כולל ירושת Roles) — מקור מועדף לבדיקות הרשאות רגישות וטבלאות מערכת.",
        "effective_privilege_grantees.csv",
        "חובה: GRANTEE, PRIVILEGE, OBJECT_TYPE. לניתוח: GRANTEE_TYPE, GRANTOR, SCHEMA_NAME, "
        "OBJECT_NAME, IS_GRANTABLE, IS_VALID",
    ),
    (
        "GRANTED_ROLES",
        "לא",
        "הקצאות תפקידים — שיוך Roles למשתמשים וזיהוי ירושת הרשאות.",
        "granted_roles.csv",
        "חובה: אחד מ-(GRANTEE / USER_NAME) ואחד מ-(ROLE_NAME / GRANTED_ROLE). "
        "לניתוח: GRANTEE_TYPE, IS_GRANTED_BY_LDAP",
    ),
    (
        "AUDIT_POLICIES",
        "כן",
        "מדיניות ניטור — האם מדיניות Audit פעילה.",
        "audit_policies.csv",
        "חובה: AUDIT_POLICY_NAME, IS_AUDIT_POLICY_ACTIVE",
    ),
    (
        "AUDIT_TRAIL",
        "לא",
        "ראיות Audit בפועל — אימות פעולות מנהליות רגישות.",
        "audit_trail.csv",
        "חובה: אחד מ-(ACTION / EVENT / STATEMENT_STRING). "
        "לניתוח: USER_NAME, TIMESTAMP / EVENT_TIME, CONNECTION_ID",
    ),
    (
        "M_INIFILE_CONTENTS",
        "כן",
        "הקשחת תצורה — ערכי INI בפועל (global.ini / indexserver.ini).",
        "m_inifile_contents.csv",
        "חובה: SECTION (או SECTION_NAME), KEY (או PARAMETER_NAME), VALUE (או CURRENT_VALUE). "
        "לניתוח: FILE_NAME",
    ),
    (
        "CONFIGURATION_PARAMETER_PROPERTIES",
        "לא",
        "מטא-דאטה פרמטרי תצורה — ברירות מחדל וטיפוסים.",
        "configuration_parameter_properties.csv",
        "חובה: SECTION, KEY, DEFAULT_VALUE. לניתוח: DESCRIPTION, DATA_TYPE_NAME, "
        "INIFILE_NAMES, RESTART_REQUIRED, IS_READ_ONLY",
    ),
]

# Detail rows: table, technical field, role, note
# Roles: חובה לקליטה | אחד-מתוך | חלופה לקליטה | ניתוח
DETAIL_ROWS: list[tuple[str, str, str, str]] = [
    # USERS
    ("USERS", "USER_NAME", "חובה לקליטה", "שם משתמש במערכת"),
    ("USERS", "LAST_SUCCESSFUL_CONNECT", "אחד-מתוך", "תאריך התחברות אחרונה — נדרש LAST_SUCCESSFUL_CONNECT או LAST_SUCCESSFUL_CONNECT_DATE"),
    ("USERS", "LAST_SUCCESSFUL_CONNECT_DATE", "אחד-מתוך", "חלופה לתאריך התחברות אחרונה"),
    ("USERS", "USER_DEACTIVATED", "ניתוח", "סטטוס נעילה / ביטול משתמש"),
    ("USERS", "CREATE_TIME", "ניתוח", "תאריך יצירת המשתמש"),
    ("USERS", "CREATOR", "ניתוח", "יוצר המשתמש"),
    ("USERS", "VALID_UNTIL", "ניתוח", "תוקף עד תאריך"),
    ("USERS", "ADMIN_GIVEN_PASSWORD", "ניתוח", "סיסמה שהוגדרה על ידי מנהל"),
    ("USERS", "IS_PASSWORD_LIFETIME_CHECK_ENABLED", "ניתוח", "האם בדיקת תוקף סיסמה פעילה (החרגות)"),
    ("USERS", "LAST_PASSWORD_CHANGE_TIME", "ניתוח", "תאריך שינוי סיסמה אחרון"),
    ("USERS", "IS_RESTRICTED", "ניתוח", "משתמש מוגבל"),
    ("USERS", "USERGROUP_NAME", "ניתוח", "קבוצת משתמשים"),
    ("USERS", "COMMENTS", "ניתוח", "הערות"),
    # M_PASSWORD_POLICY
    ("M_PASSWORD_POLICY", "PROPERTY", "חובה לקליטה", "שם פרמטר המדיניות"),
    ("M_PASSWORD_POLICY", "VALUE", "חובה לקליטה", "ערך הפרמטר"),
    # GRANTED_PRIVILEGES
    ("GRANTED_PRIVILEGES", "GRANTEE", "חובה לקליטה", "מקבל ההרשאה (משתמש / תפקיד)"),
    ("GRANTED_PRIVILEGES", "PRIVILEGE", "חובה לקליטה", "שם ההרשאה"),
    ("GRANTED_PRIVILEGES", "OBJECT_TYPE", "ניתוח", "סוג האובייקט"),
    ("GRANTED_PRIVILEGES", "GRANTEE_TYPE", "ניתוח", "סוג מקבל ההרשאה"),
    ("GRANTED_PRIVILEGES", "SCHEMA_NAME", "ניתוח", "שם סכמה"),
    ("GRANTED_PRIVILEGES", "OBJECT_NAME", "ניתוח", "שם אובייקט"),
    ("GRANTED_PRIVILEGES", "IS_GRANTABLE", "ניתוח", "האם ניתן להעביר הרשאה"),
    ("GRANTED_PRIVILEGES", "IS_VALID", "ניתוח", "האם ההרשאה תקפה"),
    # EFFECTIVE_PRIVILEGE_GRANTEES
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "GRANTEE", "חובה לקליטה", "מקבל ההרשאה האפקטיבית"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "PRIVILEGE", "חובה לקליטה", "שם ההרשאה"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "OBJECT_TYPE", "חובה לקליטה", "סוג האובייקט"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "GRANTEE_TYPE", "ניתוח", "סוג מקבל ההרשאה"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "GRANTOR", "ניתוח", "מעניק ההרשאה"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "SCHEMA_NAME", "ניתוח", "שם סכמה"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "OBJECT_NAME", "ניתוח", "שם אובייקט"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "IS_GRANTABLE", "ניתוח", "האם ניתן להעביר הרשאה"),
    ("EFFECTIVE_PRIVILEGE_GRANTEES", "IS_VALID", "ניתוח", "האם ההרשאה תקפה"),
    # GRANTED_ROLES
    ("GRANTED_ROLES", "GRANTEE", "אחד-מתוך", "מקבל התפקיד — נדרש GRANTEE או USER_NAME"),
    ("GRANTED_ROLES", "USER_NAME", "אחד-מתוך", "חלופה לשם משתמש"),
    ("GRANTED_ROLES", "ROLE_NAME", "אחד-מתוך", "שם התפקיד — נדרש ROLE_NAME או GRANTED_ROLE"),
    ("GRANTED_ROLES", "GRANTED_ROLE", "אחד-מתוך", "חלופה לשם התפקיד"),
    ("GRANTED_ROLES", "GRANTEE_TYPE", "ניתוח", "סוג מקבל התפקיד"),
    ("GRANTED_ROLES", "IS_GRANTED_BY_LDAP", "ניתוח", "הקצאה דרך LDAP"),
    # AUDIT_POLICIES
    ("AUDIT_POLICIES", "AUDIT_POLICY_NAME", "חובה לקליטה", "שם מדיניות הניטור"),
    ("AUDIT_POLICIES", "IS_AUDIT_POLICY_ACTIVE", "חובה לקליטה", "האם המדיניות פעילה"),
    # AUDIT_TRAIL
    ("AUDIT_TRAIL", "ACTION", "אחד-מתוך", "פעולה — נדרש ACTION / EVENT / STATEMENT_STRING"),
    ("AUDIT_TRAIL", "EVENT", "אחד-מתוך", "חלופה לעמודת פעולה"),
    ("AUDIT_TRAIL", "STATEMENT_STRING", "אחד-מתוך", "חלופה — טקסט פקודה / אירוע"),
    ("AUDIT_TRAIL", "USER_NAME", "ניתוח", "משתמש מבצע"),
    ("AUDIT_TRAIL", "TIMESTAMP", "ניתוח", "זמן האירוע — או EVENT_TIME"),
    ("AUDIT_TRAIL", "EVENT_TIME", "ניתוח", "חלופה לזמן האירוע"),
    ("AUDIT_TRAIL", "CONNECTION_ID", "ניתוח", "מזהה חיבור"),
    # M_INIFILE_CONTENTS
    ("M_INIFILE_CONTENTS", "SECTION", "אחד-מתוך", "סקשן — נדרש SECTION או SECTION_NAME"),
    ("M_INIFILE_CONTENTS", "SECTION_NAME", "אחד-מתוך", "חלופה לשם סקשן"),
    ("M_INIFILE_CONTENTS", "KEY", "אחד-מתוך", "מפתח פרמטר — נדרש KEY / KEY_NAME / PARAMETER_NAME"),
    ("M_INIFILE_CONTENTS", "PARAMETER_NAME", "אחד-מתוך", "חלופה לשם פרמטר"),
    ("M_INIFILE_CONTENTS", "VALUE", "אחד-מתוך", "ערך — נדרש VALUE / CONFIGURED_VALUE / CURRENT_VALUE"),
    ("M_INIFILE_CONTENTS", "CURRENT_VALUE", "אחד-מתוך", "חלופה לערך נוכחי"),
    ("M_INIFILE_CONTENTS", "FILE_NAME", "ניתוח", "שם קובץ INI (למשל global.ini)"),
    # CONFIGURATION_PARAMETER_PROPERTIES
    ("CONFIGURATION_PARAMETER_PROPERTIES", "SECTION", "חובה לקליטה", "סקשן הפרמטר"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "KEY", "חובה לקליטה", "שם הפרמטר"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "DEFAULT_VALUE", "חובה לקליטה", "ערך ברירת מחדל"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "DESCRIPTION", "ניתוח", "תיאור הפרמטר"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "DATA_TYPE_NAME", "ניתוח", "טיפוס נתונים"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "INIFILE_NAMES", "ניתוח", "קבצי INI רלוונטיים"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "RESTART_REQUIRED", "ניתוח", "האם נדרש Restart"),
    ("CONFIGURATION_PARAMETER_PROPERTIES", "IS_READ_ONLY", "ניתוח", "האם לקריאה בלבד"),
]


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def _write_data_cell(cell, value: object, *, alt: bool = False) -> None:
    cell.value = value
    cell.font = CELL_FONT
    cell.alignment = WRAP_RIGHT
    cell.border = BORDER
    if alt:
        cell.fill = ALT_FILL


def _build_summary_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "ריכוז לפי טבלה"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    headers = [
        "שם קובץ / טבלה",
        "חובה לקליטה",
        "תיאור",
        "קובץ צפוי",
        "שדות קליטה רלוונטיים",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 24

    for index, row in enumerate(SUMMARY_ROWS):
        excel_row = 2 + index
        for col, value in enumerate(row, start=1):
            _write_data_cell(ws.cell(excel_row, col), value, alt=bool(index % 2))
        ws.row_dimensions[excel_row].height = 55

    widths = [34, 14, 52, 34, 78]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:E{1 + len(SUMMARY_ROWS)}"


def _build_detail_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("פירוט שדות")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    headers = ["טבלה", "שדה טכני", "תפקיד", "הערה"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 24

    for index, row in enumerate(DETAIL_ROWS):
        excel_row = 2 + index
        for col, value in enumerate(row, start=1):
            _write_data_cell(ws.cell(excel_row, col), value, alt=bool(index % 2))
        ws.row_dimensions[excel_row].height = 28

    widths = [34, 34, 16, 70]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:D{1 + len(DETAIL_ROWS)}"


def _build_paste_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("להעתקה")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    headers = ["שם הקובץ / הסלוט", "תיאור", "שדות רלוונטיים לניתוח"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    _style_header_row(ws, 1, len(headers))

    for index, (slot, required, desc, expected, fields) in enumerate(SUMMARY_ROWS):
        excel_row = 2 + index
        values = [
            f"{slot}  |  {expected}",
            f"{desc} (חובה לקליטה: {required})",
            fields,
        ]
        for col, value in enumerate(values, start=1):
            _write_data_cell(ws.cell(excel_row, col), value, alt=bool(index % 2))
        ws.row_dimensions[excel_row].height = 60

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 78
    ws.auto_filter.ref = f"A1:C{1 + len(SUMMARY_ROWS)}"


def main() -> Path:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _build_summary_sheet(wb)
    _build_detail_sheet(wb)
    _build_paste_sheet(wb)
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    path = main()
    print(path)
